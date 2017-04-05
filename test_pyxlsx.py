# coding=utf-8
import re
import string
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell import cell


# wb = Workbook()
wb = load_workbook(filename='orders_detail.xlsx')
ws = wb.active

# freeze the row above
ws.freeze_panes = ws['A13']
# get the rows of the sheet
ROWS = ws.max_row
CLOMS = ws.max_column
print(ROWS,CLOMS)

print(ws['P12'].value)
print(type(ws['p12'].value.encode('utf-8')))
print(type(ws.cell(row=12, column=21).value.encode('utf-8')))
print(unicode('利润', 'utf-8'))

if ws['p12'].value == u'利润':  ##unicode('利润', 'utf-8') :
    print("the value is good.")


for j in range(13, ROWS+1, 1):
    count = 0
    num = 0
    pname = ''
    for i in range(17, 50, 3):
        ce = ws.cell(row=j, column=i)
        if(ce.value):
            pname= pname + str(ws.cell(row=j, column=i).value).encode('utf-8')+','
            num += int(str(ws.cell(row=j, column=i+1).value).encode('utf-8'))
            count += float(re.sub(r',','',str(ws.cell(row=j, column=i+2).value).encode('utf-8')))

            ws.cell(row=j, column=i).value = ''  #货品sku
            ws.cell(row=j, column=i+1).value = '' #货品数量
            ws.cell(row=j, column=i+2).value = '' #货品成本

        else:
            ws.cell(row = j, column = 17).value = pname
            ws.cell(row=j, column=18).value = num
            ws.cell(row=j, column=19).value = count
            break

    print("rows:"+str(j)+'    num:'+str(num)+"    count:"+str(count))
count_formula = '=SUM(' + 'S13:S' + str(ROWS) + r')'
num_formula = '=SUM(' + 'R13:R' + str(ROWS) + r')'

ws.cell(row=ROWS+1, column =17).value = ur'合计'
ws.cell(row=ROWS+1, column =18).value = num_formula
ws.cell(row=ROWS+1, column=19).value = count_formula
wb.save(filename="save.xlsx")



